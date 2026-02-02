#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
员工工时报表生成工具 - 修复版本
支持多次签到动态扩展列、斑马纹、正确的人员顺序
"""

import pandas as pd
from datetime import datetime, timedelta
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import argparse
import calendar

class ExcelReportGenerator:
    def __init__(self):
        self.raw_data = []
        self.companies = set()
        
    def parse_sheet_date(self, sheet_name, upload_date=None):
        """解析工作表名称为日期，处理跨年问题"""
        try:
            if upload_date is None:
                upload_date = datetime.now()
            
            parts = sheet_name.split('.')
            if len(parts) != 2:
                return None
                
            month = int(parts[0])
            day = int(parts[1])
            
            # 智能推断年份
            current_year = upload_date.year
            current_month = upload_date.month
            
            if current_month < month:
                year = current_year - 1
            else:
                year = current_year
                
            return datetime(year, month, day)
        except:
            return None
    
    def parse_time(self, time_value):
        """解析时间值，处理Excel时间格式问题"""
        if pd.isna(time_value) or time_value == '' or time_value is None:
            return None
            
        # 处理字符串时间格式
        if isinstance(time_value, str):
            time_value = time_value.strip()
            if ':' in time_value:
                return time_value
            return None
        
        # 处理pandas Timestamp对象 (Excel读取后的时间格式)
        if hasattr(time_value, 'strftime'):
            return time_value.strftime('%H:%M')
        
        # 处理datetime对象
        if isinstance(time_value, datetime):
            return time_value.strftime('%H:%M')
            
        # 处理Excel时间序列号
        if isinstance(time_value, (int, float)):
            try:
                if 0 <= time_value < 1:
                    total_minutes = int(time_value * 24 * 60)
                    hours = total_minutes // 60
                    minutes = total_minutes % 60
                    return f"{hours:02d}:{minutes:02d}"
                
                if time_value > 1:
                    base_date = datetime(1899, 12, 30)
                    result_datetime = base_date + timedelta(days=time_value)
                    return result_datetime.strftime('%H:%M')
                    
            except Exception as e:
                print(f"    时间解析错误: {time_value} -> {e}")
                return None
                
        return None
    
    def read_input_excel(self, file_path):
        """读取输入的Excel文件，解析所有工作表"""
        print(f"正在读取文件: {file_path}")
        
        try:
            excel_file = pd.ExcelFile(file_path)
            
            for sheet_name in excel_file.sheet_names:
                print(f"处理工作表: {sheet_name}")
                
                work_date = self.parse_sheet_date(sheet_name)
                if work_date is None:
                    print(f"跳过工作表 {sheet_name}: 无法解析日期")
                    continue
                
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=1)
                
                for idx, row in df.iterrows():
                    name_value = row.get('姓名')
                    if pd.isna(name_value) or str(name_value).strip() in ['姓名', '白班安排', '夜班安排', '']:
                        continue
                    
                    name = str(name_value).strip()
                    company = str(row.get('劳务公司', '')).strip()
                    
                    if not name or not company or company == 'nan':
                        continue
                    
                    # 解析时间
                    start_time = None
                    end_time = None
                    
                    for start_col in ['上工', '上工时间', '开始时间']:
                        if start_col in row:
                            start_time = self.parse_time(row.get(start_col))
                            break
                    
                    for end_col in ['下工', '下工时间', '结束时间']:
                        if end_col in row:
                            end_time = self.parse_time(row.get(end_col))
                            break
                    
                    self.raw_data.append({
                        'date': work_date,
                        'name': name,
                        'company': company,
                        'start_time': start_time,
                        'end_time': end_time,
                        'description': str(row.get('白班工时11H（如有延长下班的，备注原因）', '')).strip()
                    })
                    
                    self.companies.add(company)
            
            print(f"共读取 {len(self.raw_data)} 条记录")
            print(f"发现公司: {', '.join(sorted(self.companies))}")
            
        except Exception as e:
            print(f"读取文件失败: {e}")
            sys.exit(1)
    
    def generate_company_report(self, company):
        """为指定公司生成月度考勤报表"""
        company_data = [record for record in self.raw_data if record['company'] == company]
        
        if not company_data:
            return None
        
        dates = [record['date'] for record in company_data]
        min_date = min(dates)
        
        year = min_date.year
        month = min_date.month
        days_in_month = calendar.monthrange(year, month)[1]
        
        # 按首次出现顺序获取员工列表
        employees = []
        seen = set()
        for record in company_data:
            if record['name'] not in seen:
                employees.append(record['name'])
                seen.add(record['name'])
        
        # 分析每天的最大签到次数
        max_daily_records = {}
        for day in range(1, days_in_month + 1):
            max_daily_records[day] = 1  # 默认每天1次签到
        
        # 统计每天每个员工的实际签到次数
        for employee in employees:
            employee_data = [record for record in company_data if record['name'] == employee]
            daily_counts = {}
            for record in employee_data:
                day = record['date'].day
                daily_counts[day] = daily_counts.get(day, 0) + 1
            
            for day, count in daily_counts.items():
                max_daily_records[day] = max(max_daily_records.get(day, 1), count)
        
        # 创建报表数据
        report_data = []
        
        for employee in employees:
            employee_data = [record for record in company_data if record['name'] == employee]
            
            # 按日期分组
            daily_records = {}
            for record in employee_data:
                day = record['date'].day
                if day not in daily_records:
                    daily_records[day] = []
                daily_records[day].append(record)
            
            # 创建员工的上工和下工行
            start_row = {'员工姓名': employee, '类型': '上工'}
            end_row = {'员工姓名': '', '类型': '下工'}
            
            for day in range(1, days_in_month + 1):
                max_records_for_day = max_daily_records.get(day, 1)
                
                if day in daily_records:
                    records = daily_records[day]
                    for i in range(max_records_for_day):
                        col_key = f"{day}_{i+1}" if max_records_for_day > 1 else str(day)
                        if i < len(records):
                            start_row[col_key] = records[i]['start_time'] or ''
                            end_row[col_key] = records[i]['end_time'] or ''
                        else:
                            start_row[col_key] = ''
                            end_row[col_key] = ''
                else:
                    for i in range(max_records_for_day):
                        col_key = f"{day}_{i+1}" if max_records_for_day > 1 else str(day)
                        start_row[col_key] = ''
                        end_row[col_key] = ''
            
            report_data.append(start_row)
            report_data.append(end_row)
        
        return {
            'data': report_data,
            'year': year,
            'month': month,
            'days_in_month': days_in_month,
            'company': company,
            'max_daily_records': max_daily_records
        }

    def save_company_report(self, report_info, output_dir):
        """保存公司报表到Excel文件"""
        if not report_info:
            return

        company = report_info['company']
        year = report_info['year']
        month = report_info['month']
        days_in_month = report_info['days_in_month']
        data = report_info['data']
        max_daily_records = report_info['max_daily_records']

        wb = Workbook()
        ws = wb.active
        ws.title = 'sheet1'

        # 计算总列数
        total_cols = 4  # A,B,C,D 基础列
        for day in range(1, days_in_month + 1):
            total_cols += max_daily_records.get(day, 1)

        # 设置标题
        title = f"{year}年{month:02d}月"
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(name='SimSun', bold=True, size=12)

        end_col = get_column_letter(total_cols)
        ws.merge_cells(f'A1:{end_col}1')

        # 设置字体
        header_font = Font(name='SimSun', size=12)
        data_font = Font(name='SimSun', size=10)

        # 设置表头
        ws['A2'] = '序号'
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A2'].font = header_font

        ws['B2'] = '姓名/日期'
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B2'].font = header_font

        ws['C2'] = '劳务\n公司'
        ws['C2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws['C2'].font = header_font

        ws['D2'] = '上工\n时间'
        ws['D2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws['D2'].font = header_font

        ws['A3'] = ''
        ws['B3'] = ''
        ws['C3'] = ''
        ws['D3'] = '下工\n时间'
        ws['D3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws['D3'].font = header_font

        # 合并基础列
        ws.merge_cells('A2:A3')
        ws.merge_cells('B2:B3')
        ws.merge_cells('C2:C3')

        # 生成动态日期表头
        weekdays = ['一', '二', '三', '四', '五', '六', '日']
        current_col = 5  # E列开始

        for day in range(1, days_in_month + 1):
            max_records_for_day = max_daily_records.get(day, 1)

            # 计算星期几
            date_obj = datetime(year, month, day)
            weekday_idx = date_obj.weekday()
            weekday_name = weekdays[weekday_idx]

            if max_records_for_day == 1:
                # 单次签到：正常列
                col_letter = get_column_letter(current_col)
                ws[f'{col_letter}2'] = weekday_name
                ws[f'{col_letter}2'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'{col_letter}2'].font = header_font

                ws[f'{col_letter}3'] = day
                ws[f'{col_letter}3'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'{col_letter}3'].font = header_font

                current_col += 1
            else:
                # 多次签到：扩展列
                start_col = current_col
                end_col = current_col + max_records_for_day - 1

                # 合并星期行
                start_letter = get_column_letter(start_col)
                end_letter = get_column_letter(end_col)
                ws.merge_cells(f'{start_letter}2:{end_letter}2')
                ws[f'{start_letter}2'] = weekday_name
                ws[f'{start_letter}2'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'{start_letter}2'].font = header_font

                # 合并日期行
                ws.merge_cells(f'{start_letter}3:{end_letter}3')
                ws[f'{start_letter}3'] = day
                ws[f'{start_letter}3'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'{start_letter}3'].font = header_font

                current_col += max_records_for_day

        # 填充数据
        row_idx = 4
        seq_num = 1

        # 斑马纹颜色
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        for i in range(0, len(data), 2):
            if i + 1 >= len(data):
                break

            start_record = data[i]
            end_record = data[i + 1]

            employee_name = start_record.get('员工姓名', '')

            # 斑马纹：每个员工（两行）使用相同的背景色
            use_fill = (seq_num % 2 == 0)

            # 序号
            ws[f'A{row_idx}'] = seq_num
            ws.merge_cells(f'A{row_idx}:A{row_idx + 1}')
            ws[f'A{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'A{row_idx}'].font = data_font
            if use_fill:
                ws[f'A{row_idx}'].fill = light_fill

            # 姓名
            ws[f'B{row_idx}'] = employee_name
            ws.merge_cells(f'B{row_idx}:B{row_idx + 1}')
            ws[f'B{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'B{row_idx}'].font = data_font
            if use_fill:
                ws[f'B{row_idx}'].fill = light_fill

            # 劳务公司
            ws[f'C{row_idx}'] = company
            ws.merge_cells(f'C{row_idx}:C{row_idx + 1}')
            ws[f'C{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'C{row_idx}'].font = data_font
            if use_fill:
                ws[f'C{row_idx}'].fill = light_fill

            # 上工/下工标识
            ws[f'D{row_idx}'] = '上工'
            ws[f'D{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'D{row_idx}'].font = data_font
            if use_fill:
                ws[f'D{row_idx}'].fill = light_fill

            ws[f'D{row_idx + 1}'] = '下工'
            ws[f'D{row_idx + 1}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'D{row_idx + 1}'].font = data_font
            if use_fill:
                ws[f'D{row_idx + 1}'].fill = light_fill

            # 填充时间数据
            current_col = 5
            for day in range(1, days_in_month + 1):
                max_records_for_day = max_daily_records.get(day, 1)

                for record_idx in range(max_records_for_day):
                    col_letter = get_column_letter(current_col)
                    col_key = f"{day}_{record_idx+1}" if max_records_for_day > 1 else str(day)

                    # 上工时间
                    start_time = start_record.get(col_key, '')
                    ws[f'{col_letter}{row_idx}'] = start_time
                    ws[f'{col_letter}{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws[f'{col_letter}{row_idx}'].font = data_font
                    if use_fill:
                        ws[f'{col_letter}{row_idx}'].fill = light_fill

                    # 下工时间
                    end_time = end_record.get(col_key, '')
                    ws[f'{col_letter}{row_idx + 1}'] = end_time
                    ws[f'{col_letter}{row_idx + 1}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws[f'{col_letter}{row_idx + 1}'].font = data_font
                    if use_fill:
                        ws[f'{col_letter}{row_idx + 1}'].fill = light_fill

                    current_col += 1

            row_idx += 2
            seq_num += 1

        # 添加边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in range(1, row_idx):
            for col in range(1, total_cols + 1):
                ws.cell(row=row, column=col).border = thin_border

        # 冻结表头（前3行）
        ws.freeze_panes = 'A4'  # 冻结A1:A3行，从第4行开始可滚动

        # 保存文件（添加月份信息）
        filename = f"employee_hours-{month:02d}-{company}.xlsx"
        filepath = os.path.join(output_dir, filename)
        wb.save(filepath)
        print(f"已生成报表: {filepath}")

        return filepath

def main():
    """主函数"""
    parser = argparse.ArgumentParser(description='员工工时报表生成工具')
    parser.add_argument('input_file', help='输入的Excel文件路径')
    parser.add_argument('-o', '--output', default='.', help='输出目录 (默认: 当前目录)')

    args = parser.parse_args()

    if not os.path.exists(args.input_file):
        print(f"错误: 输入文件不存在: {args.input_file}")
        sys.exit(1)

    output_dir = args.output
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    generator = ExcelReportGenerator()
    generator.read_input_excel(args.input_file)

    if not generator.raw_data:
        print("错误: 没有读取到有效数据")
        sys.exit(1)

    generated_files = []
    for company in sorted(generator.companies):
        print(f"\n正在生成 {company} 的报表...")
        report_info = generator.generate_company_report(company)
        if report_info:
            filepath = generator.save_company_report(report_info, output_dir)
            generated_files.append(filepath)

    print(f"\n✅ 报表生成完成!")
    print(f"共生成 {len(generated_files)} 个文件:")
    for filepath in generated_files:
        print(f"  - {filepath}")

if __name__ == "__main__":
    main()
