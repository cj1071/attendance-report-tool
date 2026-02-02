#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
考勤统计报表生成脚本 - 直接从原始签到表生成
"""

import os
import sys
from excel_report_generator_fixed import ExcelReportGenerator
from attendance_calculator import AttendanceCalculator
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def generate_attendance_stats(input_file, output_dir=None):
    """
    从原始签到表直接生成考勤统计
    
    Args:
        input_file: 原始签到表文件路径
        output_dir: 输出目录
    """
    if output_dir is None:
        output_dir = os.getcwd()
    
    print("🚀 考勤统计报表生成器")
    print("=" * 80)
    print(f"📄 处理文件: {os.path.basename(input_file)}")
    
    # 第一步：读取原始数据（使用工时报表生成器的逻辑）
    generator = ExcelReportGenerator()
    calculator = AttendanceCalculator()
    
    print("\n📖 正在读取Excel文件...")
    generator.read_input_excel(input_file)
    
    if not generator.raw_data:
        print("❌ 没有读取到有效数据")
        return
    
    print(f"共读取 {len(generator.raw_data)} 条记录")
    print(f"发现公司: {', '.join(sorted(generator.companies))}")
    
    # 第二步：为每个公司生成考勤统计报表
    print(f"\n📊 开始生成考勤统计报表...")
    
    for company in sorted(generator.companies):
        print(f"  正在生成 {company} 的考勤统计...")
        
        # 筛选该公司的数据
        company_data = [rec for rec in generator.raw_data if rec.get('company') == company]
        
        print(f"    {company} 的记录数: {len(company_data)}")
        
        if not company_data:
            print(f"    跳过 {company}（没有数据）")
            continue
        
        # 从数据中提取年月（参考 employee_hours 逻辑）
        dates = [rec['date'] for rec in company_data]
        min_date = min(dates)
        year = min_date.year
        month = min_date.month
        
        # 计算考勤统计
        statistics = []
        for rec in company_data:
            # 将字段映射为 calculator 期望的格式
            mapped_rec = {
                '姓名': rec.get('name'),
                '劳务公司': rec.get('company'),
                '上工时间': rec.get('start_time'),
                '下工时间': rec.get('end_time')
            }
            
            stat = calculator.process_attendance_record(mapped_rec)
            
            # 添加年月日信息
            from datetime import datetime
            date_obj = rec.get('date')
            if isinstance(date_obj, datetime):
                stat.update({
                    'year': date_obj.year,
                    'month': date_obj.month,
                    'day': date_obj.day
                })
            else:
                # 如果不是datetime对象，尝试从字符串解析
                stat.update({
                    'year': rec.get('year'),
                    'month': rec.get('month'),
                    'day': rec.get('day')
                })
            
            statistics.append(stat)
        
        print(f"    计算了 {len(statistics)} 条统计数据")
        
        # 生成报表
        output_file = os.path.join(output_dir, f"attendance_stats-{month:02d}-{company}.xlsx")
        
        try:
            generate_excel_report(statistics, output_file)
            print(f"  ✓ {os.path.basename(output_file)}")
        except Exception as e:
            print(f"  ✗ 生成失败: {e}")
            import traceback
            traceback.print_exc()
    
    print(f"\n✅ 考勤统计生成完成!")

def generate_excel_report(statistics, output_file):
    """生成Excel考勤统计报表"""
    
    if not statistics:
        return
    
    # 按员工和日期分组
    employee_stats = {}
    employee_order = []
    
    for stat in statistics:
        key = (stat['name'], stat['company'])
        if key not in employee_stats:
            employee_stats[key] = {}
            employee_order.append(key)
        
        day = stat['day']
        if day not in employee_stats[key]:
            employee_stats[key][day] = []
        
        employee_stats[key][day].append(stat)
    
    # 计算每天最多的签到次数
    max_checkins_per_day = {}
    for day in range(1, 32):
        max_checkins_per_day[day] = 0
        for daily_stats in employee_stats.values():
            if day in daily_stats:
                max_checkins_per_day[day] = max(max_checkins_per_day[day], len(daily_stats[day]))
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "考勤统计"
    
    year = statistics[0]['year']
    month = statistics[0]['month']
    
    # 样式
    header_font = Font(name='SimSun', size=12, bold=True)
    data_font = Font(name='SimSun', size=10)
    title_font = Font(name='SimSun', size=14, bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    night_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    zebra_fill_1 = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    zebra_fill_2 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # 第1行：标题
    ws.merge_cells('A1:C1')
    cell = ws.cell(row=1, column=1, value=f"{year}年{month:02d}月")
    cell.font = title_font
    cell.alignment = left_align
    
    # 第3行：表头
    ws.cell(row=3, column=1, value='序号').font = header_font
    ws.cell(row=3, column=1).alignment = center_align
    ws.cell(row=3, column=1).border = border
    ws.cell(row=3, column=1).fill = header_fill
    
    ws.cell(row=3, column=2, value='姓名/日期').font = header_font
    ws.cell(row=3, column=2).alignment = center_align
    ws.cell(row=3, column=2).border = border
    ws.cell(row=3, column=2).fill = header_fill
    
    ws.cell(row=3, column=3, value='劳务公司').font = header_font
    ws.cell(row=3, column=3).alignment = center_align
    ws.cell(row=3, column=3).border = border
    ws.cell(row=3, column=3).fill = header_fill
    
    # 日期列（合并表头）
    col_idx = 4
    for day in range(1, 32):
        if max_checkins_per_day[day] > 0:
            start_col = col_idx
            checkins = max_checkins_per_day[day]
            
            # 如果该天有多次签到，合并表头
            if checkins > 1:
                end_col = start_col + checkins - 1
                ws.merge_cells(start_row=3, start_column=start_col, 
                             end_row=3, end_column=end_col)
            
            # 设置表头内容和样式
            cell = ws.cell(row=3, column=start_col, value=f"{day}日")
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            cell.fill = header_fill
            
            # 为合并的单元格也设置边框
            for i in range(checkins):
                cell = ws.cell(row=3, column=start_col + i)
                cell.border = border
                cell.fill = header_fill
            
            col_idx += checkins
    
    # 汇总列（增加出勤次数）
    total_col = col_idx
    for i, title in enumerate(['出勤次数', '总工时', '夜班补贴次数', '夜班补贴']):
        cell = ws.cell(row=3, column=total_col + i, value=title)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        cell.fill = header_fill
    
    # 设置列宽
    ws.column_dimensions['A'].width = 6   # 序号
    ws.column_dimensions['B'].width = 12  # 姓名
    ws.column_dimensions['C'].width = 12  # 劳务公司
    for col in range(4, total_col):
        ws.column_dimensions[get_column_letter(col)].width = 6  # 日期列
    ws.column_dimensions[get_column_letter(total_col)].width = 10      # 出勤次数
    ws.column_dimensions[get_column_letter(total_col + 1)].width = 10  # 总工时
    ws.column_dimensions[get_column_letter(total_col + 2)].width = 12  # 夜班补贴次数
    ws.column_dimensions[get_column_letter(total_col + 3)].width = 10  # 夜班补贴
    
    # 写入数据
    row_idx = 4
    seq_num = 1
    
    for (name, company) in employee_order:
        daily_stats = employee_stats[(name, company)]
        zebra_fill = zebra_fill_1 if seq_num % 2 == 0 else zebra_fill_2
        
        # 序号、姓名、公司
        ws.cell(row=row_idx, column=1, value=seq_num).font = data_font
        ws.cell(row=row_idx, column=1).alignment = center_align
        ws.cell(row=row_idx, column=1).border = border
        ws.cell(row=row_idx, column=1).fill = zebra_fill
        
        ws.cell(row=row_idx, column=2, value=name).font = data_font
        ws.cell(row=row_idx, column=2).alignment = center_align
        ws.cell(row=row_idx, column=2).border = border
        ws.cell(row=row_idx, column=2).fill = zebra_fill
        
        ws.cell(row=row_idx, column=3, value=company).font = data_font
        ws.cell(row=row_idx, column=3).alignment = center_align
        ws.cell(row=row_idx, column=3).border = border
        ws.cell(row=row_idx, column=3).fill = zebra_fill
        
        # 统计数据
        total_hours = 0.0
        night_allowance_count = 0
        total_night_allowance = 0.0
        attendance_days = 0  # 出勤次数（一天算一次，不管几次签到）
        
        # 填充每天数据
        col_idx = 4
        for day in range(1, 32):
            if max_checkins_per_day[day] > 0:
                if day in daily_stats:
                    # 统计出勤天数（一天有签到就算一天）
                    attendance_days += 1
                    
                    for stat in daily_stats[day]:
                        value = stat['effective_hours']
                        total_hours += value
                        
                        if stat['night_allowance'] > 0:
                            night_allowance_count += 1
                            total_night_allowance += stat['night_allowance']
                        
                        cell = ws.cell(row=row_idx, column=col_idx, value=round(value, 1))
                        cell.font = data_font
                        cell.alignment = center_align
                        cell.border = border
                        cell.fill = night_fill if stat['is_night_shift'] else zebra_fill
                        col_idx += 1
                    
                    # 填充空列
                    for _ in range(len(daily_stats[day]), max_checkins_per_day[day]):
                        cell = ws.cell(row=row_idx, column=col_idx, value='')
                        cell.border = border
                        cell.fill = zebra_fill
                        col_idx += 1
                else:
                    for _ in range(max_checkins_per_day[day]):
                        cell = ws.cell(row=row_idx, column=col_idx, value='')
                        cell.border = border
                        cell.fill = zebra_fill
                        col_idx += 1
        
        # 汇总列（新增出勤次数）
        # 出勤次数
        ws.cell(row=row_idx, column=total_col, value=attendance_days).font = data_font
        ws.cell(row=row_idx, column=total_col).alignment = center_align
        ws.cell(row=row_idx, column=total_col).border = border
        ws.cell(row=row_idx, column=total_col).fill = zebra_fill
        
        # 总工时
        ws.cell(row=row_idx, column=total_col + 1, value=round(total_hours, 1)).font = data_font
        ws.cell(row=row_idx, column=total_col + 1).alignment = center_align
        ws.cell(row=row_idx, column=total_col + 1).border = border
        ws.cell(row=row_idx, column=total_col + 1).fill = zebra_fill
        
        # 夜班补贴次数
        ws.cell(row=row_idx, column=total_col + 2, value=night_allowance_count).font = data_font
        ws.cell(row=row_idx, column=total_col + 2).alignment = center_align
        ws.cell(row=row_idx, column=total_col + 2).border = border
        ws.cell(row=row_idx, column=total_col + 2).fill = zebra_fill
        
        # 夜班补贴金额
        ws.cell(row=row_idx, column=total_col + 3, value=round(total_night_allowance, 1)).font = data_font
        ws.cell(row=row_idx, column=total_col + 3).alignment = center_align
        ws.cell(row=row_idx, column=total_col + 3).border = border
        ws.cell(row=row_idx, column=total_col + 3).fill = zebra_fill
        
        row_idx += 1
        seq_num += 1
    
    wb.save(output_file)

if __name__ == "__main__":
    # 查找输入文件
    current_dir = os.getcwd()
    input_files = [f for f in os.listdir(current_dir) if '劳务签到表' in f and f.endswith('.xls')]
    
    if not input_files:
        print("❌ 当前目录没有找到劳务签到表文件")
        sys.exit(1)
    
    input_file = os.path.join(current_dir, input_files[0])
    generate_attendance_stats(input_file, current_dir)

